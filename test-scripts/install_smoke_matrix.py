#!/usr/bin/env python3
"""Install pinned packages in batches and run smoke checks in uv and pepip modes."""

from __future__ import annotations

import argparse
import os
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence

from _smoke_matrix_utils import delete_dir
from rich.console import Console
from rich.panel import Panel
from rich.prompt import IntPrompt
from rich.table import Table

CONSOLE = Console()
PYTHON_VERSION = "3.10"
VENV_DIRNAME = ".venv"

DEFAULT_PINNED_VERSIONS = {
    "numpy": "2.1.3",
    "pandas": "2.2.3",
    "requests": "2.32.3",
    "scipy": "1.14.1",
    "matplotlib": "3.9.2",
    "scikit-learn": "1.5.2",
    "sqlalchemy": "2.0.36",
    "fastapi": "0.115.4",
    "click": "8.1.7",
    "pydantic": "2.9.2",
    "httpx": "0.27.2",
    "rich": "13.9.4",
    "polars": "1.12.0",
    "pillow": "11.0.0",
    "pyyaml": "6.0.2",
    "jinja2": "3.1.4",
    "aiohttp": "3.10.10",
    "sympy": "1.13.3",
    "openpyxl": "3.1.5",
    "networkx": "3.4.2",
    "tqdm": "4.66.6",
    "beautifulsoup4": "4.12.3",
    "lxml": "5.3.0",
    "orjson": "3.10.11",
    "python-dateutil": "2.9.0.post0",
    "uvicorn": "0.32.0",
}

SMOKE_CODE_BY_PACKAGE = {
    "numpy": r"""
import numpy as np
arr = np.array([1, 2, 3, 4], dtype=np.float64)
assert np.isclose(arr.mean(), 2.5)
""",
    "pandas": r"""
import pandas as pd
frame = pd.DataFrame({"x": [1, 2], "y": [10, 20]})
assert frame["y"].sum() == 30
""",
    "requests": r"""
import requests
req = requests.Request("GET", "https://example.com")
prepared = req.prepare()
assert prepared.method == "GET"
""",
    "scipy": r"""
import numpy as np
from scipy import linalg
mat = np.array([[2.0, 0.0], [0.0, 5.0]])
inv = linalg.inv(mat)
assert np.allclose(inv, np.array([[0.5, 0.0], [0.0, 0.2]]))
""",
    "matplotlib": r"""
import matplotlib
assert matplotlib.__version__
""",
    "scikit-learn": r"""
import numpy as np
from sklearn.linear_model import LinearRegression
X = np.array([[1], [2], [3], [4]], dtype=np.float64)
y = np.array([2, 4, 6, 8], dtype=np.float64)
model = LinearRegression().fit(X, y)
assert np.isclose(model.coef_[0], 2.0)
""",
    "sqlalchemy": r"""
from sqlalchemy import create_engine, text
engine = create_engine("sqlite:///:memory:")
with engine.connect() as conn:
    value = conn.execute(text("SELECT 42")).scalar_one()
assert value == 42
""",
    "fastapi": r"""
import fastapi
app = fastapi.FastAPI()

@app.get("/ping")
def ping():
    return {"ok": True}

assert any(route.path == "/ping" for route in app.routes)
""",
    "click": r"""
import click

@click.command()
def cmd():
    pass

assert cmd.name == "cmd"
""",
    "pydantic": r"""
import pydantic

class User(pydantic.BaseModel):
    name: str

u = User(name="alice")
assert u.name == "alice"
""",
    "httpx": r"""
import httpx
client = httpx.Client()
request = client.build_request("GET", "https://example.com")
assert request.method == "GET"
client.close()
""",
    "rich": r"""
from rich.console import Console
console = Console(record=True)
console.print("ok")
assert "ok" in console.export_text()
""",
    "polars": r"""
import polars as pl
df = pl.DataFrame({"a": [1, 2, 3]})
assert df.shape == (3, 1)
""",
    "pillow": r"""
from PIL import Image
img = Image.new("RGB", (10, 10))
assert img.size == (10, 10)
""",
    "pyyaml": r"""
import yaml
obj = yaml.safe_load("a: 1")
assert obj["a"] == 1
""",
    "jinja2": r"""
from jinja2 import Template
out = Template("Hello {{ name }}").render(name="world")
assert out == "Hello world"
""",
    "aiohttp": r"""
import aiohttp
assert aiohttp.__version__
""",
    "sympy": r"""
import sympy
x = sympy.Symbol("x")
expr = sympy.expand((x + 1) ** 2)
assert str(expr) == "x**2 + 2*x + 1"
""",
    "openpyxl": r"""
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws["A1"] = "ok"
assert ws["A1"].value == "ok"
""",
    "networkx": r"""
import networkx as nx
G = nx.Graph()
G.add_edge("a", "b")
assert nx.has_path(G, "a", "b")
""",
    "tqdm": r"""
from tqdm import tqdm
assert tqdm is not None
""",
    "beautifulsoup4": r"""
from bs4 import BeautifulSoup
soup = BeautifulSoup("<html><body><p>ok</p></body></html>", "html.parser")
assert soup.p.text == "ok"
""",
    "lxml": r"""
from lxml import etree
root = etree.fromstring(b"<root><child>ok</child></root>")
assert root.findtext("child") == "ok"
""",
    "orjson": r"""
import orjson
data = orjson.loads(orjson.dumps({"x": 1}))
assert data["x"] == 1
""",
    "python-dateutil": r"""
from dateutil.parser import parse
dt = parse("2026-01-01")
assert dt.year == 2026
""",
    "uvicorn": r"""
import uvicorn
assert uvicorn.__version__
""",
}


def run(
    cmd: List[str], cwd: Path, env: Optional[Dict[str, str]] = None
) -> subprocess.CompletedProcess:
    ui_command(cmd)
    completed = subprocess.run(
        cmd,
        cwd=str(cwd),
        env=env,
        text=True,
        capture_output=True,
        check=False,
    )
    if completed.returncode != 0:
        raise RuntimeError(
            f"Command failed ({completed.returncode}): {' '.join(cmd)}\n"
            f"stdout:\n{completed.stdout}\n"
            f"stderr:\n{completed.stderr}"
        )
    return completed


def ensure_binary(name: str) -> None:
    if shutil.which(name) is None:
        raise FileNotFoundError(f"Required executable not found in PATH: {name}")


def venv_path(work_dir: Path) -> Path:
    return work_dir / VENV_DIRNAME


def venv_python(work_dir: Path) -> Path:
    scripts_dir = "Scripts" if os.name == "nt" else "bin"
    executable = "python.exe" if os.name == "nt" else "python"
    return venv_path(work_dir) / scripts_dir / executable


def venv_env(work_dir: Path) -> Dict[str, str]:
    env = os.environ.copy()
    scripts_dir = "Scripts" if os.name == "nt" else "bin"
    bin_dir = venv_path(work_dir) / scripts_dir
    env["VIRTUAL_ENV"] = str(venv_path(work_dir))
    env["PATH"] = f"{bin_dir}{os.pathsep}{env.get('PATH', '')}"
    return env


def reset_venv(work_dir: Path, selected_mode: str) -> None:
    target = venv_path(work_dir)

    if target.exists() or target.is_symlink():
        ui_header("Removing existing virtual environment", str(target))

        if target.is_dir() and not target.is_symlink():
            delete_dir(target)
        else:
            target.unlink()

    creation_mode = "uv" if selected_mode in {"uv", "all"} else "pepip"
    ui_header(
        "Creating virtual environment",
        f"mode={creation_mode}; path={target}",
    )

    if creation_mode == "uv":
        ensure_binary("uv")
        run(
            ["uv", "venv", "--clear", "--python", PYTHON_VERSION, str(target)],
            cwd=work_dir,
        )
    elif creation_mode == "pepip":
        run([sys.executable, "-m", "venv", str(target)], cwd=work_dir)
    else:
        raise ValueError(
            f"Unsupported virtual environment creation mode: {creation_mode}"
        )


def package_batches(
    packages: Sequence[str],
    batch_size: int,
    start_index: int = 0,
    limit: Optional[int] = None,
) -> Iterable[List[str]]:
    if batch_size < 1:
        raise ValueError("--batch-size must be greater than or equal to 1.")
    if start_index < 0:
        raise ValueError("--start must be greater than or equal to 0.")
    if start_index > len(packages):
        raise ValueError(
            "--start must be less than or equal to the package count "
            f"({len(packages)})."
        )
    if limit is not None and limit < 0:
        raise ValueError("--limit must be greater than or equal to 0.")

    selected_packages = packages[start_index:]
    if limit is not None:
        selected_packages = selected_packages[:limit]

    for start in range(0, len(selected_packages), batch_size):
        yield list(selected_packages[start : start + batch_size])


def package_name(spec: str) -> str:
    for separator in ("===", "==", ">=", "<=", "~=", "!=", ">", "<"):
        if separator in spec:
            return spec.split(separator, 1)[0].strip()
    return spec.strip()


def smoke_code_for_batch(batch: Sequence[str]) -> str:
    chunks = []
    missing = []

    for spec in batch:
        name = package_name(spec)
        smoke_code = SMOKE_CODE_BY_PACKAGE.get(name)
        if smoke_code is None:
            missing.append(name)
        else:
            chunks.append(f"# --- {name} ---\n{smoke_code.strip()}")

    if missing:
        raise KeyError(
            "No smoke code mapped for package(s): " + ", ".join(sorted(set(missing)))
        )

    return "\n\n".join(chunks)


def run_smoke_code(work_dir: Path, batch: Sequence[str], batch_idx: int) -> None:
    ui_header(f"Running smoke batch {batch_idx}", ", ".join(batch))
    run(
        [str(venv_python(work_dir)), "-c", smoke_code_for_batch(batch)],
        cwd=work_dir,
        env=venv_env(work_dir),
    )


def install_and_smoke_batches_uv(
    work_dir: Path,
    packages: Sequence[str],
    batch_size: int,
    start_index: int,
    limit: Optional[int],
) -> None:
    for idx, batch in enumerate(
        package_batches(packages, batch_size, start_index, limit), start=1
    ):
        ui_header(f"Installing uv batch {idx}", ", ".join(batch))
        run(
            ["uv", "pip", "install", "--python", str(venv_python(work_dir)), *batch],
            cwd=work_dir,
        )
        run_smoke_code(work_dir, batch, idx)


def install_and_smoke_batches_pepip(
    work_dir: Path,
    packages: Sequence[str],
    batch_size: int,
    start_index: int,
    limit: Optional[int],
) -> None:
    for idx, batch in enumerate(
        package_batches(packages, batch_size, start_index, limit), start=1
    ):
        ui_header(f"Installing pepip batch {idx}", ", ".join(batch))
        run(
            [sys.executable, "-m", "pepip.cli", "install", *batch],
            cwd=work_dir,
            env=venv_env(work_dir),
        )
        run_smoke_code(work_dir, batch, idx)


def run_uv_mode(
    work_dir: Path,
    packages: Sequence[str],
    batch_size: int,
    start_index: int,
    limit: Optional[int],
) -> None:
    install_and_smoke_batches_uv(work_dir, packages, batch_size, start_index, limit)


def run_pepip_mode(
    work_dir: Path,
    packages: Sequence[str],
    batch_size: int,
    start_index: int,
    limit: Optional[int],
) -> None:
    install_and_smoke_batches_pepip(work_dir, packages, batch_size, start_index, limit)


AVAILABLE_MODES = ("uv", "pepip")


def modes_to_run(requested: str) -> Iterable[str]:
    if requested == "all":
        return AVAILABLE_MODES
    return (requested,)


def ui_print(message: str, *, style: Optional[str] = None) -> None:
    CONSOLE.print(message, style=style)


def ui_command(cmd: Sequence[str]) -> None:
    ui_print(f"$ {' '.join(cmd)}", style="dim")


def ui_header(title: str, subtitle: str = "") -> None:
    text = title if not subtitle else f"{title}\n[dim]{subtitle}[/dim]"
    CONSOLE.print(Panel.fit(text, border_style="cyan"))


def ui_packages(packages: Sequence[str]) -> None:
    table = Table(title="Pinned packages", show_lines=False)
    table.add_column("#", justify="right")
    table.add_column("Package")
    table.add_column("Version")
    for idx, spec in enumerate(packages, start=1):
        name, _, version = spec.partition("==")
        table.add_row(str(idx), name, version or "custom")
    CONSOLE.print(table)


def select_mode_interactively() -> str:
    options = ["all", *AVAILABLE_MODES]
    table = Table(title="Select install mode")
    table.add_column("Option", justify="right")
    table.add_column("Mode")
    table.add_column("Description")
    descriptions = {
        "all": "Run uv and pepip",
        "uv": "Use uv-managed .venv + uv pip",
        "pepip": "Use pepip CLI against ./.venv",
    }
    for idx, mode in enumerate(options, start=1):
        table.add_row(str(idx), mode, descriptions[mode])

    CONSOLE.print(table)
    choice = IntPrompt.ask(
        "Choose an option",
        choices=[str(i) for i in range(1, len(options) + 1)],
        default=1,
    )
    return options[choice - 1]


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Install pinned popular packages in batches and run smoke operations "
            "using uv or pepip. Uses ./.venv in the current working directory."
        )
    )
    parser.add_argument(
        "--mode",
        choices=["uv", "pepip", "all"],
        default=None,
        help="Which install mode to run. If omitted, an interactive menu is shown.",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=5,
        help="Number of packages to install per batch.",
    )
    parser.add_argument(
        "--start",
        type=int,
        default=0,
        help=(
            "Zero-based package index to start from. This is independent of "
            "--batch-size; e.g. --start 7 starts at package #8."
        ),
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help=(
            "Maximum number of packages to process after applying --start. "
            "This is independent of --batch-size."
        ),
    )
    parser.add_argument(
        "--package",
        action="append",
        default=[],
        help=(
            "Override package list with explicit pinned specifiers. "
            "Provide multiple times, e.g. --package numpy==2.1.3."
        ),
    )
    args = parser.parse_args()

    packages = (
        args.package
        if args.package
        else [f"{name}=={version}" for name, version in DEFAULT_PINNED_VERSIONS.items()]
    )
    if not packages:
        raise ValueError("No packages specified.")
    if args.start < 0:
        raise ValueError("--start must be greater than or equal to 0.")
    if args.start > len(packages):
        raise ValueError(
            f"--start must be less than or equal to the package count "
            f"({len(packages)})."
        )
    if args.limit is not None and args.limit < 0:
        raise ValueError("--limit must be greater than or equal to 0.")

    selected_mode = args.mode or select_mode_interactively()
    selected_modes = tuple(modes_to_run(selected_mode))
    work_dir = Path.cwd()

    ui_header(
        "Package smoke runner",
        (
            f"Selected mode: {selected_mode}; batch size: {args.batch_size}; "
            f"start index: {args.start}; limit: {args.limit}; "
            f"python: {PYTHON_VERSION}; workdir={work_dir}"
        ),
    )
    ui_packages(packages)
    reset_venv(work_dir, selected_mode)

    for mode in selected_modes:
        ui_header(f"Running mode: {mode}", f"workdir={work_dir}")
        if mode == "uv":
            run_uv_mode(work_dir, packages, args.batch_size, args.start, args.limit)
        elif mode == "pepip":
            run_pepip_mode(work_dir, packages, args.batch_size, args.start, args.limit)
        else:
            raise ValueError(f"Unsupported mode: {mode}")

    ui_print("\nAll selected modes passed.", style="bold green")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
